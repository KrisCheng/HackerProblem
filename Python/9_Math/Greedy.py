from datetime import datetime
import dateutil.parser
from openpyxl import load_workbook
from matplotlib import pyplot
import pylab as plt
from collections import Counter

WIDE_BODY = ["332", "333", "33E", "33H", "33L", "773"]
NARROW_BODY = ["319", "320", "321", "323", "325", "738", "73A", "73E", "73H", "73L"]

# 5分钟为时间步长
TIME_STEPS = 288 # 60/5 * 12

def __time_transfer(begin_time, end_time, priority):
    begin = begin_time.split(":")
    end = end_time.split(":")
    if(len(begin) == 3):
        begin.pop()
    if(len(end) == 3):
        end.pop()
    begin_index = int(int(begin[0].strip())*12) + int(int(begin[1].strip())/5)
    end_index = int(int(end[0].strip())*12) + int(int(end[1].strip())/5)
    if priority == 1:
        begin_index = 0
    elif priority == 3:
        end_index = 287
    dict = {}
    dict["begin_index"] = begin_index
    dict["end_index"] = end_index
    return dict

def __fitness_function(pucks_list):
    num_satisfy_airline = 0
    for puck in pucks_list:
        if puck["是否分配"] == 1:
            if(puck["优先级"] == 2):
                num_satisfy_airline = num_satisfy_airline + 2
            else:
                num_satisfy_airline = num_satisfy_airline + 1
    return num_satisfy_airline

def get_num_empty_gates(gates_list):
    num_empty_gates = 0
    for gate in gates_list:
        if sum(gate["资源数组"]) == 0:
            num_empty_gates = num_empty_gates + 1
    return num_empty_gates

def __get_earliest_time(candidate_list, gates_list):
    best_gate = ""
    tmp_time = TIME_STEPS
    for candidate in candidate_list:
        for gate in gates_list:
            if gate["登机口"] == candidate:
                try:
                    result = Counter(gate["资源数组"])
                    index = gate["资源数组"].index(1)
                    last_time = result[1] + index
                    if last_time < tmp_time:
                        tmp_time = last_time
                        best_gate = gate["登机口"]
                except:
                    best_gate = gate["登机口"]
                finally:
                    break
    return best_gate

def __load_data_sources():
    file = "InputData.xlsx"
    list_pucks = []
    list_tickets = []
    list_gates = []

    wb = load_workbook(filename=file, data_only=True, read_only=True)

    # Pucks
    ws_pucks = wb.worksheets[0]
    # Tickets
    ws_tickets = wb.worksheets[1]
    # Gates
    ws_gates = wb.worksheets[2]

    idx1 = [str(cell.value).replace('\n','') for cell in ws_pucks[1]]
    idx2 = [str(cell.value).replace('\n','') for cell in ws_tickets[1]]
    idx3 = [str(cell.value).replace('\n','') for cell in ws_gates[1]]

    for row in ws_pucks.iter_rows(row_offset=1):
        dict = {}
        for cell in row:
            cell_value = cell.value
            if cell_value is not None:
                cell_key = idx1[cell.column - 1]
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime("%Y-%m-%d")
                dict[cell_key] = str(cell_value)
        # 筛除不需要考虑的航班
        if not(dict["到达日期"] != "2018-01-20" and dict["出发日期"] != "2018-01-20"):
            list_pucks.append(dict)

    for row in ws_tickets.iter_rows(row_offset=1):
        dict = {}
        for cell in row:
            cell_value = cell.value
            if cell_value is not None:
                cell_key = idx2[cell.column - 1]
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime("%Y-%m-%d")
                dict[cell_key] = str(cell_value)
        if not(dict["到达日期"] != "2018-01-20"):
            list_tickets.append(dict)

    for row in ws_gates.iter_rows(row_offset=1):
        dict = {}
        for cell in row:
            cell_value = cell.value
            if cell_value is not None:
                cell_key = idx3[cell.column - 1]
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime("%Y-%m-%d")
                dict[cell_key] = str(cell_value)
        list_gates.append(dict)

    return list_pucks, list_tickets, list_gates

def main_task():
    list_pucks, list_tickets, list_gates = __load_data_sources()
    
    # 2. 优先级处理 考虑--“每架飞机转场的到达和出发两个航班必须分配在同一登机口进行，其间不能挪移别处；”
    for puck in list_pucks:
    # 飞机分类
    # 19号到达，20号起飞 优先级 1
    # 20号到达，20号起飞 优先级 2
    # 19号到达，20号起飞 优先级 3
        if (puck["到达日期"] == '2018-01-19' and puck["出发日期"] == '2018-01-20'):
            puck["优先级"] = 1
        elif (puck["到达日期"] == '2018-01-20' and puck["出发日期"] == '2018-01-20'):
            puck["优先级"] = 2 
        elif (puck["到达日期"] == '2018-01-20' and puck["出发日期"] == '2018-01-21'):
            puck["优先级"] = 3
        else:
            print("Error!!!")

    # 添加机体类别
    for puck in list_pucks:
        if puck["飞机型号"] in WIDE_BODY:
            puck["机体类别"] = "W"
        elif puck["飞机型号"] in NARROW_BODY:
            puck["机体类别"] = "N"
        else:
            print("Error！！！")

    # 根据优先级对飞机排序
    tmp_list = []
    first_count = 1
    for puck in list_pucks:
        if puck["优先级"] == 1:
            tmp_list.insert(1, puck)
            first_count = first_count + 1
        if puck["优先级"] == 3:
            tmp_list.append(puck)
        if puck["优先级"] == 2:
            tmp_list.insert(first_count, puck)
    list_pucks = tmp_list
    
    # 3. 同一优先级中采用FIFO, 冒泡排序, 获取最终飞机分配优先队列
    for i in range(len(list_pucks) - 1):
        for j in range(len(list_pucks) - i - 1):
            if list_pucks[j]["到达时刻"] > list_pucks[j+1]["到达时刻"] and list_pucks[j]["优先级"] == list_pucks[j+1]["优先级"]:
                list_pucks[j], list_pucks[j+1] = list_pucks[j+1], list_pucks[j]

    for gate in list_gates:
        resource_time = []
        for i in range(TIME_STEPS):
            resource_time.append(0)
        gate["资源数组"] = resource_time
    
    # 贪心策略
    for puck in list_pucks:
        puck["是否分配"] = 0
        puck["对应登机口"] = "NONE"
        resource_period = __time_transfer(puck["到达时刻"], puck["出发时刻"], puck["优先级"])
        puck["到达时刻"] = resource_period["begin_index"]
        puck["出发时刻"] = resource_period["end_index"]

        candidate_list = []
        for gate in list_gates: 
            # 判断飞机型号 / 到达和出发类型是否 匹配
            if (puck["机体类别"].strip() != gate["机体类别"].strip()
             or puck["到达类型"].strip() not in gate["到达类型"]
             or puck["出发类型"].strip() not in gate["出发类型"]):
                continue
            # 判断时间上是否冲突
            time_conflict = False
            for i in range(resource_period["begin_index"], resource_period["end_index"]+1):
                if gate["资源数组"][i] != 0:
                    time_conflict = True
                    break
            if time_conflict:
                continue
            # 获得一个候选解
            candidate_list.append(gate["登机口"])
        
        if candidate_list is not []:
            # 找到离降落时间最近的点
            best_gate = __get_earliest_time(candidate_list, list_gates)
            for i in range(resource_period["begin_index"], resource_period["end_index"]+1):
                for gate in list_gates:
                    if gate["登机口"] == best_gate:
                        puck["是否分配"] = 1
                        gate["资源数组"][i] = 1
                        puck["对应登机口"] = gate["登机口"]
                        # 间隔延迟45分钟 9个break
                        if (resource_period["end_index"] < TIME_STEPS - 9):
                            for i in range(1,10):
                                gate["资源数组"][resource_period["end_index"]+i] = 1
                        else:
                            for i in range(resource_period["end_index"], TIME_STEPS):
                                gate["资源数组"][i] = 1
                        break

    # list_pucks 和 list_gates 分别表示当前的最优情况
    for gate in list_gates:
        if sum(gate["资源数组"]) == 0:
            gate["是否为空"] = 1
        else:
            gate["是否为空"] = 0
            
    list_satisfy_airline = []
    list_unsatisfy_airline = []
    
    num_all_airline = 0

    num_satisfy_airline = 0
    num_satisfy_airline_wide = 0
    num_satisfy_airline_narrow = 0

    num_free_gate = 0
    num_free_gate_narrow = 0
    num_free_gate_wide = 0

    gate_resource_narrow = []
    gate_resource_wide = []
    list_free_gate = []

    for puck in list_pucks:
        if puck["优先级"] == 2:
            num_all_airline = num_all_airline + 2
        else:
            num_all_airline = num_all_airline + 1
        if puck["是否分配"] == 1:
            list_satisfy_airline.append(puck)
            if puck["优先级"] == 2:
                # 每架飞机匹配两个航班
                num_satisfy_airline = num_satisfy_airline + 2
                if puck["机体类别"] == "N":
                    num_satisfy_airline_narrow = num_satisfy_airline_narrow + 2
                elif puck["机体类别"] == "W":
                    num_satisfy_airline_wide = num_satisfy_airline_wide + 2
            elif puck["优先级"] == 1 or puck["优先级"] == 3:
                # 每架飞机匹配一个航班
                num_satisfy_airline = num_satisfy_airline + 1
                if puck["机体类别"] == "N":
                    num_satisfy_airline_narrow = num_satisfy_airline_narrow + 1
                elif puck["机体类别"] == "W":
                    num_satisfy_airline_wide = num_satisfy_airline_wide + 1
        if puck["是否分配"] == 0:
            list_unsatisfy_airline.append(puck)

    print("num_satisfy_airline : %s " % num_satisfy_airline)

    for gate in list_gates:
        if(gate["机体类别"] == "N"):
            gate_resource_narrow.append(gate["资源数组"])
        if(gate["机体类别"] == "W"):
            gate_resource_wide.append(gate["资源数组"])
        if sum(gate["资源数组"]) == 0:
            num_free_gate = num_free_gate + 1
            if(gate["机体类别"] == "N"):
                num_free_gate_narrow = num_free_gate_narrow + 1
            if(gate["机体类别"] == "W"):
                num_free_gate_wide = num_free_gate_wide + 1            
            list_free_gate.append(gate)
            
    fig = plt.figure(figsize=(20, 4))
    ax = fig.add_subplot(221)
    plt.imshow(gate_resource_wide)
    ax.text(1, 2, 'Wide', fontsize=12, color='r')
    cbar = plt.colorbar(plt.imshow(gate_resource_wide), orientation='horizontal')

    ax = fig.add_subplot(222)
    plt.imshow(gate_resource_narrow)
    ax.text(1, 2, 'Narrow', fontsize=12, color='r')

    # plt.title('Resource Ratio', fontsize=12)
    cbar = plt.colorbar(plt.imshow(gate_resource_narrow), orientation='horizontal')
    cbar.set_label('0-1',fontsize=12)
    pyplot.show()

    print("num_all_airline : %s " % num_all_airline)
    print("num_satisfy_airline : %s " % num_satisfy_airline)
    print("num_satisfy_airline_narrow : %s " % num_satisfy_airline_narrow)
    print("num_satisfy_airline_wide : %s " % num_satisfy_airline_wide)
    # for satisfy_airline in list_satisfy_airline:
    #     print(satisfy_airline)
    # for unsatisfy_airline in list_unsatisfy_airline:
    #     print(unsatisfy_airline)

    print("---")
    print("num_free_gate : %s " % num_free_gate)
    print("num_free_gate_narrow : %s " % num_free_gate_narrow)
    print("num_free_gate_wide : %s " % num_free_gate_wide)
    # for free_gate in list_free_gate:
    #     print(free_gate)
main_task()
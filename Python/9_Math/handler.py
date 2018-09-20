result_list = []
from openpyxl import load_workbook
from datetime import datetime
import dateutil.parser
file=open("Q3_result.txt")
lines=file.readlines()
for line in lines:
    d = eval(line)
    result_list.append(d)

file = "final.xlsx"
wb = load_workbook(filename=file, data_only=True, read_only=False)
ws_pucks = wb.worksheets[0]
for row in ws_pucks.iter_rows(row_offset=1):
    test = row[0].value
    for i in range(len(result_list)):
        if (str(result_list[i]["飞机转场记录号"]).strip() == str(test).strip()
            and result_list[i]["对应登机口"] != "NONE"):
            row[14].value = result_list[i]["对应登机口"]
wb.save('final.xlsx')

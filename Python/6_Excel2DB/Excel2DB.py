#!/usr/bin/python
# -*- coding: utf-8 -*-

'''
Desc: create a excel, read data from it, and save them to the datatbase, remain to refactoring.
Author: Kris Peng
Copyright (c) 2018 - Kris Peng <kris.dacpc@gmail.com>
'''
from openpyxl import load_workbook
from pandas import DataFrame
from sqlalchemy import Column, String, create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
import numpy as np

# 创建对象的基类:
Base = declarative_base()

class Account(Base):
    __tablename__ = 'account'

    id = Column(String(20), primary_key=True)
    name = Column(String(50))
    email = Column(String(50))

dest_filename = 'workbook.xlsx'
wb = load_workbook(dest_filename)
ws = wb.get_sheet_by_name('MySheet')
list = []

for columns in ws.iter_rows(row_offset = 0):
    for column in columns:
        list.append(str(column.value))
list = np.array(list)
list = list.reshape([4,3])
list = list[1:]
# DB part
engine = create_engine('mysql+pymysql://root:pengcheng00@localhost:3306/test?charset=utf8mb4')
DBSession = sessionmaker(bind=engine)
# 创建session对象:
session = DBSession()
# 创建新User对象:
for row in list:
    new_Account = Account(id = row[0], name = row[1], email = row[2])
    # 添加到session:
    session.add(new_Account)
# 提交即保存到数据库:
session.commit()
# 关闭session:
session.close()